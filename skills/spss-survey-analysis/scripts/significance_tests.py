#!/usr/bin/env python3
"""
Run chi-square significance tests and add results to an Excel workbook.

Usage:
    significance_tests.py <sav_file> [options]

Options:
    --banners VAR1,VAR2     Banner variables to test against
    --rows VAR1,VAR2        Row variables (default: all Likert + recoded)
    --output <file.xlsx>    Excel file to add sheet to (or create new)

Examples:
    significance_tests.py outputs/survey_recoded.sav --banners Country,Age,Gender --output outputs/banner_tables.xlsx
"""

import sys
from pathlib import Path

try:
    import pyreadstat
    import pandas as pd
    import numpy as np
    from scipy.stats import chi2_contingency
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing packages. Run: pip install pyreadstat pandas numpy scipy openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from classify_variables import classify_all

# Color fills
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")   # p < 0.05
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")   # p < 0.10
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)


def run_chi_square(df, row_var, banner_var):
    """Run chi-square test for one row variable against one banner variable."""
    try:
        ct = pd.crosstab(df[row_var].dropna(), df[banner_var].dropna())
        if ct.shape[0] < 2 or ct.shape[1] < 2:
            return None
        chi2, p, dof, expected = chi2_contingency(ct)
        return {
            "row_variable": row_var,
            "banner_variable": banner_var,
            "chi_square": round(float(chi2), 4),
            "p_value": round(float(p), 6),
            "dof": int(dof),
            "significant_05": p < 0.05,
            "significant_10": p < 0.10
        }
    except Exception as e:
        return {
            "row_variable": row_var,
            "banner_variable": banner_var,
            "error": str(e)
        }


def run_all_tests(df, row_vars, banner_vars):
    """Run chi-square for all row × banner combinations."""
    results = []
    for row_var in row_vars:
        if row_var not in df.columns:
            continue
        for ban_var in banner_vars:
            if ban_var not in df.columns:
                continue
            result = run_chi_square(df, row_var, ban_var)
            if result:
                results.append(result)
    return results


def write_significance_sheet(wb, results):
    """Add or replace a Significance Tests sheet in the workbook."""
    sheet_name = "Significance Tests"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    headers = ["Row Variable", "Banner Variable", "Chi-Square", "p-value", "df", "Significant (p<.05)", "Marginal (p<.10)"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for i, r in enumerate(results):
        row = i + 2
        ws.cell(row=row, column=1, value=r.get("row_variable", ""))
        ws.cell(row=row, column=2, value=r.get("banner_variable", ""))

        if "error" in r:
            ws.cell(row=row, column=3, value=f"Error: {r['error']}")
            continue

        ws.cell(row=row, column=3, value=r["chi_square"])
        p_cell = ws.cell(row=row, column=4, value=r["p_value"])
        p_cell.number_format = "0.000000"
        ws.cell(row=row, column=5, value=r["dof"])

        sig_05 = ws.cell(row=row, column=6, value="Yes" if r["significant_05"] else "No")
        sig_10 = ws.cell(row=row, column=7, value="Yes" if r["significant_10"] else "No")

        # Color coding
        if r["significant_05"]:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = GREEN_FILL
        elif r["significant_10"]:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ORANGE_FILL

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    return ws


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    sav_path = Path(sys.argv[1])
    banner_vars = []
    specific_rows = None
    output_path = None

    args = sys.argv[2:]
    i = 0
    while i < len(args):
        if args[i] == "--banners" and i + 1 < len(args):
            banner_vars = [v.strip() for v in args[i + 1].split(",")]
            i += 2
        elif args[i] == "--rows" and i + 1 < len(args):
            specific_rows = [v.strip() for v in args[i + 1].split(",")]
            i += 2
        elif args[i] == "--output" and i + 1 < len(args):
            output_path = Path(args[i + 1])
            i += 2
        else:
            i += 1

    if not sav_path.exists():
        print(f"File not found: {sav_path}")
        sys.exit(1)

    print(f"Loading {sav_path}...")
    df, meta = pyreadstat.read_sav(str(sav_path))

    # Auto-detect if needed
    if not banner_vars:
        results = classify_all(sav_path)
        banner_vars = results["summary"]["banner_candidates"][:5]

    if specific_rows:
        row_vars = specific_rows
    else:
        results = classify_all(sav_path)
        likert = results["summary"]["likert"]
        recoded = [c for c in df.columns if any(c.startswith(v + "_top") or c.startswith(v + "_bottom") for v in likert)]
        row_vars = likert + recoded

    print(f"Testing {len(row_vars)} row vars × {len(banner_vars)} banner vars...")
    test_results = run_all_tests(df, row_vars, banner_vars)

    sig_count = sum(1 for r in test_results if r.get("significant_05", False))
    marginal_count = sum(1 for r in test_results if r.get("significant_10", False) and not r.get("significant_05", False))
    print(f"  {len(test_results)} tests run")
    print(f"  {sig_count} significant (p<.05), {marginal_count} marginal (p<.10)")

    # Write to Excel
    if output_path and output_path.exists():
        wb = load_workbook(str(output_path))
    else:
        wb = Workbook()
        if output_path is None:
            output_path = Path("outputs/significance_tests.xlsx")

    write_significance_sheet(wb, test_results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nSaved to: {output_path} (sheet: 'Significance Tests')")


if __name__ == "__main__":
    main()
