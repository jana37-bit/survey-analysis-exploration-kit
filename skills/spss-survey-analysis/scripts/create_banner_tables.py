#!/usr/bin/env python3
"""
Generate Excel banner tables from SPSS data.

Creates a formatted Excel workbook with per-variable blocks showing:
  - Variable name (bold)
  - Question text (italic)
  - Base row (n per banner column for this specific variable)
  - Value rows with PERCENTAGES (not counts)
  - Separator

Includes BOTH original full distributions AND recoded T2B variables.

Usage:
    create_banner_tables.py <sav_file> <output.xlsx> [options]

Options:
    --banners VAR1,VAR2         Banner column variables
    --rows VAR1,VAR2            Specific row variables (default: auto-detect)
    --skip-empty-banners        Omit banner columns with 0 respondents
    --include-counts            Add count columns alongside percentages

Examples:
    create_banner_tables.py outputs/survey_recoded.sav outputs/banner_tables.xlsx --banners Country,Gender,Age
    create_banner_tables.py data/survey.sav outputs/tables.xlsx --banners Region --skip-empty-banners
"""

import sys
from pathlib import Path

try:
    import pyreadstat
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing packages. Run: pip install pyreadstat pandas numpy openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from classify_variables import classify_all


# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
VARNAME_FONT = Font(name="Arial", bold=True, size=11, color="000000")
LABEL_FONT = Font(name="Arial", italic=True, size=10, color="666666")
BASE_FONT = Font(name="Arial", italic=True, size=10, color="444444")
DATA_FONT = Font(name="Arial", size=10)
DASH_FONT = Font(name="Arial", size=10, color="999999")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="2F5496")


def build_banner_structure(df, meta, banner_vars, skip_empty=False):
    """
    Build banner column definitions.
    Returns list of dicts: {name, filter, base_n}
    """
    columns = [{"name": "Total", "filter": None, "base_n": len(df)}]

    for bvar in banner_vars:
        val_labels = meta.variable_value_labels.get(bvar, {})
        if val_labels:
            values = sorted(val_labels.keys())
        else:
            values = sorted(df[bvar].dropna().unique())

        for val in values:
            label = val_labels.get(val, str(val)) if val_labels else str(val)
            n = int((df[bvar] == val).sum())

            if skip_empty and n == 0:
                continue

            columns.append({
                "name": f"{bvar}: {label}",
                "filter": (bvar, val),
                "base_n": n
            })

    return columns


def compute_percentage(df, row_var, row_val, banner_filter):
    """Compute column percentage for one cell."""
    if banner_filter is None:
        base = df[row_var].notna().sum()
        count = ((df[row_var] == row_val) & df[row_var].notna()).sum()
    else:
        bvar, bval = banner_filter
        mask = (df[bvar] == bval)
        base = (mask & df[row_var].notna()).sum()
        count = (mask & (df[row_var] == row_val)).sum()

    if base == 0:
        return None
    return round(float(count / base * 100), 1)


def compute_var_base(df, row_var, banner_filter):
    """Compute valid-n base for a variable within a banner group."""
    if banner_filter is None:
        return int(df[row_var].notna().sum())
    else:
        bvar, bval = banner_filter
        return int(((df[bvar] == bval) & df[row_var].notna()).sum())


def write_header_row(ws, banner_columns):
    """Write and style the frozen header row."""
    cell = ws.cell(row=1, column=1, value="")
    cell.fill = HEADER_FILL

    for j, col in enumerate(banner_columns):
        cell = ws.cell(row=1, column=j + 2, value=col["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER

    ws.freeze_panes = "B2"
    return 2


def write_section_header(ws, row, text, n_data_cols):
    """Write a colored section header spanning all columns."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for j in range(n_data_cols):
        ws.cell(row=row, column=j + 2).fill = SECTION_FILL
    return row + 2  # extra blank row after section header


def write_variable_block(ws, current_row, df, meta, var, banner_columns,
                         is_recoded=False, recoded_label_override=None,
                         recoded_val_labels_override=None):
    """
    Write one complete variable block.

    Structure:
      Row: Variable name (bold)
      Row: Question text (italic gray)
      Row: Base (n) — per banner column, for THIS variable
      Row: Value 1 — percentage%
      Row: Value 2 — percentage%
      ...
      Row: thin border separator
      Row: blank
    """
    # ── Resolve metadata ─────────────────────────────────────────────
    val_labels = meta.variable_value_labels.get(var, {})
    var_label = ""

    if var in meta.column_names:
        idx = list(meta.column_names).index(var)
        if meta.column_labels and idx < len(meta.column_labels):
            var_label = str(meta.column_labels[idx])

    # Override for recoded variables
    if recoded_label_override:
        var_label = recoded_label_override
    if recoded_val_labels_override:
        val_labels = recoded_val_labels_override

    # For recoded vars without explicit labels, try to derive
    if not var_label and is_recoded:
        for suffix in ['_top2', '_top3', '_T2B', '_bottom2']:
            if var.endswith(suffix):
                orig_name = var[:-len(suffix)]
                if orig_name in meta.column_names:
                    orig_idx = list(meta.column_names).index(orig_name)
                    if meta.column_labels and orig_idx < len(meta.column_labels):
                        box_type = suffix.replace('_', '').replace('top', 'Top ').replace('bottom', 'Bottom ').replace('T2B', 'Top 2 Box')
                        var_label = f"{meta.column_labels[orig_idx]} [{box_type.strip()}]"
                break

    if not val_labels and is_recoded:
        val_labels = {0.0: "Bottom box", 1.0: "Top 2 Box"}

    if var not in df.columns:
        return current_row

    # ── Row 1: Variable name (bold) ──────────────────────────────────
    ws.cell(row=current_row, column=1, value=var).font = VARNAME_FONT
    current_row += 1

    # ── Row 2: Question text (italic) ────────────────────────────────
    if var_label:
        cell = ws.cell(row=current_row, column=1, value=var_label)
        cell.font = LABEL_FONT
        cell.alignment = LEFT
        current_row += 1

    # ── Row 3: Base row ──────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="Base (n)").font = BASE_FONT
    for j, col in enumerate(banner_columns):
        base = compute_var_base(df, var, col["filter"])
        cell = ws.cell(row=current_row, column=j + 2, value=base)
        cell.font = BASE_FONT
        cell.alignment = CENTER
        cell.number_format = '#,##0'
    current_row += 1

    # ── Value rows with percentages ──────────────────────────────────
    if val_labels:
        value_order = sorted(val_labels.keys())
    else:
        vals = df[var].dropna().unique()
        value_order = sorted([v for v in vals if not np.isnan(v)] if len(vals) > 0 else [])

    for val in value_order:
        label = str(val_labels.get(val, val)) if val_labels else str(val)
        cell = ws.cell(row=current_row, column=1, value=f"  {label}")
        cell.font = DATA_FONT
        cell.alignment = LEFT

        for j, col in enumerate(banner_columns):
            pct = compute_percentage(df, var, val, col["filter"])
            cell = ws.cell(row=current_row, column=j + 2)

            if pct is None:
                cell.value = "-"
                cell.font = DASH_FONT
            else:
                cell.value = pct / 100  # Store as decimal for Excel % format
                cell.number_format = '0.0%'  # Excel renders as "76.5%"
            cell.alignment = CENTER

        current_row += 1

    # ── Separator ────────────────────────────────────────────────────
    for j in range(len(banner_columns) + 1):
        ws.cell(row=current_row, column=j + 1).border = THIN_BORDER
    current_row += 2  # blank row after separator

    return current_row


def determine_row_variables(df, meta):
    """
    Auto-detect row variables: include BOTH original and recoded.

    Returns list of (var_name, is_recoded) tuples.
    Each original variable is followed by its recoded counterpart(s).
    """
    recoded_suffixes = ['_top2', '_top3', '_T2B', '_bottom2']
    recoded_vars = set(c for c in df.columns if any(c.endswith(s) for s in recoded_suffixes))

    # Map originals → their recoded versions
    original_to_recoded = {}
    for rec in recoded_vars:
        for suffix in recoded_suffixes:
            if rec.endswith(suffix):
                orig = rec[:-len(suffix)]
                if orig in df.columns:
                    original_to_recoded.setdefault(orig, []).append(rec)
                break

    row_vars = []
    seen = set()

    # Original + recoded pairs (interleaved)
    for col in df.columns:
        if col in original_to_recoded and col not in seen:
            row_vars.append((col, False))
            seen.add(col)
            for rec in sorted(original_to_recoded[col]):
                row_vars.append((rec, True))
                seen.add(rec)

    # Orphaned recoded vars (no matching original)
    for rec in sorted(recoded_vars - seen):
        row_vars.append((rec, True))

    return row_vars


def create_banner_tables(sav_path, output_path, banner_vars, specific_rows=None,
                         skip_empty=False):
    """Main function: create the complete banner tables workbook."""
    print(f"Loading {sav_path}...")
    df, meta = pyreadstat.read_sav(str(sav_path))
    print(f"  {meta.number_rows} respondents, {meta.number_columns} variables")

    # Auto-detect banners
    if not banner_vars:
        results = classify_all(sav_path)
        banner_vars = results["summary"]["banner_candidates"][:5]
        print(f"  Auto-detected banners: {', '.join(banner_vars)}")

    # Build banner structure
    banner_columns = build_banner_structure(df, meta, banner_vars, skip_empty)
    n_data_cols = len(banner_columns)
    print(f"  Banner columns: {n_data_cols} (Total + {n_data_cols - 1} groups)")

    if skip_empty:
        all_vals = sum(len(meta.variable_value_labels.get(bvar, {})) or int(df[bvar].nunique())
                       for bvar in banner_vars)
        skipped = all_vals - (n_data_cols - 1)
        if skipped > 0:
            print(f"  Skipped {skipped} empty banner columns (0 respondents)")

    # Determine row variables
    if specific_rows:
        row_vars = [(v, v.endswith(('_top2', '_top3', '_T2B', '_bottom2'))) for v in specific_rows if v in df.columns]
    else:
        row_vars = determine_row_variables(df, meta)

    n_orig = sum(1 for _, is_rec in row_vars if not is_rec)
    n_rec = sum(1 for _, is_rec in row_vars if is_rec)
    print(f"  Row variables: {len(row_vars)} ({n_orig} original full distributions + {n_rec} recoded T2B)")

    if n_orig == 0 and n_rec > 0:
        print(f"\n  ⚠ WARNING: No original variables found alongside recoded versions.")
        print(f"  The table will only show T2B recoded values, not the full scale distributions.")
        print(f"  To fix this, ensure the recoded .sav file contains BOTH original AND _top2 variables.\n")

    # ── Build workbook ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Banner Tables"

    current_row = write_header_row(ws, banner_columns)

    # Write all variable blocks
    last_was_recoded = None
    for var, is_recoded in row_vars:
        # Insert section headers at transitions
        if last_was_recoded is None and not is_recoded and n_rec > 0:
            current_row = write_section_header(
                ws, current_row,
                "ORIGINAL VARIABLES — Full Scale Distributions",
                n_data_cols
            )
        elif is_recoded and last_was_recoded is False and n_orig > 0:
            current_row = write_section_header(
                ws, current_row,
                "RECODED VARIABLES — Top 2 Box",
                n_data_cols
            )

        current_row = write_variable_block(
            ws, current_row, df, meta, var, banner_columns, is_recoded
        )
        last_was_recoded = is_recoded

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 55
    for j in range(n_data_cols):
        ws.column_dimensions[get_column_letter(j + 2)].width = 18

    # Autofilter
    last_col = get_column_letter(n_data_cols + 1)
    ws.auto_filter.ref = f"A1:{last_col}1"

    # ── Save ─────────────────────────────────────────────────────────
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"\n✅ Saved: {output_path}")
    print(f"  Rows: {current_row - 1} | Columns: {n_data_cols + 1}")
    print(f"  Original distributions: {n_orig} variables with all scale points")
    print(f"  Recoded T2B: {n_rec} variables")
    print(f"  Per-variable base rows: ✓")
    print(f"  Values formatted as: percentages (e.g., 76.5%)")

    return wb


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    sav_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    banner_vars = []
    specific_rows = None
    skip_empty = False

    args = sys.argv[3:]
    i = 0
    while i < len(args):
        if args[i] == "--banners" and i + 1 < len(args):
            banner_vars = [v.strip() for v in args[i + 1].split(",")]
            i += 2
        elif args[i] == "--rows" and i + 1 < len(args):
            specific_rows = [v.strip() for v in args[i + 1].split(",")]
            i += 2
        elif args[i] == "--skip-empty-banners":
            skip_empty = True
            i += 1
        else:
            i += 1

    if not sav_path.exists():
        print(f"File not found: {sav_path}")
        sys.exit(1)

    create_banner_tables(sav_path, output_path, banner_vars, specific_rows, skip_empty)


if __name__ == "__main__":
    main()
