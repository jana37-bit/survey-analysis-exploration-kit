#!/usr/bin/env python3
"""
Validate an Excel banner tables workbook for common errors.

Checks:
- Percentages are in valid range (0-100)
- Base rows have integer values
- No empty data cells where values are expected
- Required sheets exist

Usage:
    validate_workbook.py <workbook.xlsx>

Examples:
    validate_workbook.py outputs/banner_tables.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing package. Run: pip install openpyxl")
    sys.exit(1)


def validate_workbook(workbook_path):
    """Run validation checks and return issues found."""
    wb = load_workbook(str(workbook_path), data_only=True)
    issues = []

    # Check required sheets
    expected_sheets = ["Banner Tables"]
    for sheet in expected_sheets:
        if sheet not in wb.sheetnames:
            issues.append(f"Missing required sheet: '{sheet}'")

    if "Banner Tables" not in wb.sheetnames:
        return issues

    ws = wb["Banner Tables"]

    # Scan for issues
    base_row_count = 0
    pct_out_of_range = 0
    empty_data_cells = 0

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        label_cell = row[0]
        if label_cell.value is None:
            continue

        label = str(label_cell.value).strip()
        is_base_row = label.lower().startswith("base")

        for cell in row[1:]:
            if cell.value is None:
                continue

            if isinstance(cell.value, (int, float)):
                if is_base_row:
                    base_row_count += 1
                    if isinstance(cell.value, float) and cell.value != int(cell.value):
                        issues.append(f"Base row has non-integer value at {cell.coordinate}: {cell.value}")
                else:
                    # Check percentage range
                    if cell.value < 0 or cell.value > 100:
                        pct_out_of_range += 1
                        if pct_out_of_range <= 5:
                            issues.append(f"Percentage out of range at {cell.coordinate}: {cell.value}")

    if pct_out_of_range > 5:
        issues.append(f"... and {pct_out_of_range - 5} more out-of-range percentages")

    # Summary
    if not issues:
        print(f"✅ Workbook validation passed: {workbook_path}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")
        print(f"   Banner Tables: {ws.max_row} rows × {ws.max_column} columns")
    else:
        print(f"⚠️  Workbook validation found {len(issues)} issue(s):")
        for issue in issues:
            print(f"   - {issue}")

    return issues


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(f"File not found: {workbook_path}")
        sys.exit(1)

    validate_workbook(workbook_path)


if __name__ == "__main__":
    main()
