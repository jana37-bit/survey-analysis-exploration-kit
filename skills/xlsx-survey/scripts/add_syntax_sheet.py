#!/usr/bin/env python3
"""
Add SPSS syntax content as a sheet in an existing Excel workbook.

Usage:
    add_syntax_sheet.py <workbook.xlsx> <syntax.sps>

Examples:
    add_syntax_sheet.py outputs/banner_tables.xlsx outputs/spss_custom_tables_syntax.sps
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Missing package. Run: pip install openpyxl")
    sys.exit(1)


def add_syntax_sheet(workbook_path, syntax_path):
    """Add SPSS syntax as a sheet in the workbook."""
    wb = load_workbook(str(workbook_path))

    sheet_name = "SPSS Syntax"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    with open(syntax_path, "r") as f:
        lines = f.readlines()

    mono_font = Font(name="Consolas", size=10)
    comment_font = Font(name="Consolas", size=10, color="008000")

    for i, line in enumerate(lines):
        cell = ws.cell(row=i + 1, column=1, value=line.rstrip())
        if line.strip().startswith("*"):
            cell.font = comment_font
        else:
            cell.font = mono_font

    ws.column_dimensions["A"].width = 100
    wb.save(str(workbook_path))
    print(f"Added '{sheet_name}' sheet to {workbook_path}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    workbook_path = Path(sys.argv[1])
    syntax_path = Path(sys.argv[2])

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        sys.exit(1)
    if not syntax_path.exists():
        print(f"Syntax file not found: {syntax_path}")
        sys.exit(1)

    add_syntax_sheet(workbook_path, syntax_path)


if __name__ == "__main__":
    main()
